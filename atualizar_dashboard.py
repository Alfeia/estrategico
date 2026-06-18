# -*- coding: utf-8 -*-
"""
Automatização da base do dashboard de Risco Estratégico (Farol).

Fluxo executado a cada mês:
  1. Localiza, dentro de uma pasta, a planilha de KRI mais recente
     (nomes no padrão "01. Janeiro", "02. Fevereiro", ...).
  2. Lê as 4 abas de risco e extrai a nota final de cada risco
     (coluna "Nota Final"; na aba de Desequilíbrio Atuarial a coluna
     chama-se "EPS (KRI)"). O valor usado é o da última linha
     preenchida da coluna — a linha-resumo "KRI", de 0% a 100%.
  3. Preenche a aba "Fato_Indicadores" da planilha destino seguindo o
     padrão existente: Data, ID_Risco, Código, Valor realizado,
     Meta (90%), Variação e Status (farol).
  4. Calcula a variação comparando o mês atual com o mês anterior do
     mesmo risco.
  5. Define o farol: Verde >= 90%; Amarelo > 70% e < 90%; Vermelho < 70%.
  6. Atualiza a aba "Dim_Tempo" com o mês/data processado.

Uso:
    python atualizar_dashboard.py
ou, sobrescrevendo parâmetros:
    python atualizar_dashboard.py --pasta KRIs --destino "dados farol ficticios_8710.xlsx" --ano 2026

Basta colocar a nova planilha de KRIs na pasta e rodar o script.
"""

from __future__ import annotations

import argparse
import calendar
import datetime as dt
import re
import shutil
import sys
import unicodedata
from pathlib import Path

import openpyxl

# --------------------------------------------------------------------------- #
# Configuração padrão (pode ser sobrescrita por linha de comando)
# --------------------------------------------------------------------------- #
PASTA_KRI_PADRAO = "KRIs"                              # pasta das planilhas de KRI
DESTINO_PADRAO = "dados farol ficticios_8710.xlsx"    # base do dashboard
META_PADRAO = 0.9                                      # meta de atingimento (90%)

ABA_FATO = "Fato_Indicadores"
ABA_TEMPO = "Dim_Tempo"
ABA_RISCO = "Dim_Risco"

# Abreviações de mês usadas na aba Dim_Tempo (padrão já existente).
MES_NOME = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}
MES_PT_PARA_NUM = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5,
    "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}


# --------------------------------------------------------------------------- #
# Utilitários
# --------------------------------------------------------------------------- #
def _normalizar(texto) -> str:
    """Minúsculas, sem acentos e com espaços colapsados (para comparar nomes)."""
    if texto is None:
        return ""
    texto = unicodedata.normalize("NFKD", str(texto))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().lower()


def _eh_numero(valor) -> bool:
    return isinstance(valor, (int, float)) and not isinstance(valor, bool)


def status_farol(valor: float) -> str:
    """Classifica o farol conforme a regra do negócio.

    Verde    -> valor >= 90%
    Amarelo  -> valor > 70% e < 90%
    Vermelho -> valor <= 70%
    """
    if valor >= 0.90:
        return "Verde"
    if valor > 0.70:
        return "Amarelo"
    return "Vermelho"


def trimestre(data: dt.date) -> str:
    return f"{data.year}-T{(data.month - 1) // 3 + 1}"


def fim_do_mes(ano: int, mes: int) -> dt.datetime:
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    return dt.datetime(ano, mes, ultimo_dia)


# --------------------------------------------------------------------------- #
# 1. Localizar a planilha de KRI mais recente
# --------------------------------------------------------------------------- #
def _mes_do_nome(nome_arquivo: str):
    """Extrai o número do mês a partir do nome do arquivo.

    Reconhece tanto o prefixo numérico ("01. Janeiro") quanto o nome do
    mês por extenso ("Indicadores janeiro").
    """
    base = Path(nome_arquivo).stem
    m = re.match(r"\s*(\d{1,2})\s*[.\-_) ]", base)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return num
    norm = _normalizar(base)
    for nome, num in MES_PT_PARA_NUM.items():
        if nome in norm:
            return num
    return None


def localizar_planilha_recente(pasta: Path) -> tuple[Path, int]:
    """Retorna (caminho, numero_do_mes) da planilha de KRI mais recente."""
    if not pasta.is_dir():
        raise FileNotFoundError(f"Pasta de KRIs não encontrada: {pasta}")

    candidatos = []
    for arq in pasta.iterdir():
        if arq.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        if arq.name.startswith("~$"):  # arquivos temporários do Excel
            continue
        mes = _mes_do_nome(arq.name)
        if mes is not None:
            candidatos.append((mes, arq.stat().st_mtime, arq))

    if not candidatos:
        raise FileNotFoundError(
            f"Nenhuma planilha de KRI reconhecida na pasta: {pasta}"
        )

    # Mais recente = maior número de mês; desempate pela data de modificação.
    candidatos.sort(key=lambda t: (t[0], t[1]))
    mes, _mtime, caminho = candidatos[-1]
    return caminho, mes


# --------------------------------------------------------------------------- #
# 2. Ler as abas de risco e extrair a nota final de cada risco
# --------------------------------------------------------------------------- #
# Cabeçalhos procurados, em ordem de preferência.
CABECALHOS = ["nota final", "eps (kri)"]


def _linha_cabecalho(ws):
    """Linha de cabeçalho das abas de risco.

    Primeiro procura a linha que contém 'Fator de Risco'. Se o modelo
    mudar, cai para uma linha que tenha o cabeçalho da nota ('Nota Final'
    ou 'EPS (KRI)') junto de vários outros cabeçalhos — assim não confunde
    com o texto "Nota Final" que aparece como rótulo de linha (sozinho) na
    coluna de fatores.
    """
    limite = min(ws.max_row, 15)
    for r in range(1, limite + 1):
        for c in range(1, ws.max_column + 1):
            if _normalizar(ws.cell(row=r, column=c).value) == "fator de risco":
                return r
    # Fallback: linha com cabeçalho de nota e pelo menos 5 células preenchidas.
    for r in range(1, limite + 1):
        textos = [_normalizar(ws.cell(row=r, column=c).value)
                  for c in range(1, ws.max_column + 1)]
        preenchidas = sum(1 for t in textos if t)
        if preenchidas >= 5 and any(t in CABECALHOS for t in textos):
            return r
    return None


def _localizar_coluna(ws):
    """Localiza (linha_cabecalho, coluna) do cabeçalho de nota do risco.

    A busca é feita apenas na linha de cabeçalho (a que contém 'Fator de
    Risco'), para não confundir com o texto "Nota Final" que também
    aparece como rótulo de linha na coluna de fatores.

    Preferimos "Nota Final"; se não existir (caso do Desequilíbrio
    Atuarial), usamos "EPS (KRI)". Quando o cabeçalho aparece em mais de
    uma coluna, usamos a última (que contém a nota consolidada do risco).
    """
    linha = _linha_cabecalho(ws)
    if linha is None:
        return None
    for cab in CABECALHOS:
        colunas = [c for c in range(1, ws.max_column + 1)
                   if _normalizar(ws.cell(row=linha, column=c).value) == cab]
        if colunas:
            return linha, colunas[-1]  # última coluna correspondente
    return None


def extrair_nota(ws) -> float:
    """Última nota numérica preenchida na coluna de nota do risco."""
    pos = _localizar_coluna(ws)
    if pos is None:
        raise ValueError(
            f"Aba '{ws.title}': não encontrei coluna 'Nota Final' nem 'EPS (KRI)'."
        )
    linha_cab, col = pos
    valor = None
    for r in range(linha_cab + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if _eh_numero(v):
            valor = float(v)
    if valor is None:
        raise ValueError(f"Aba '{ws.title}': nenhuma nota numérica encontrada.")
    return valor


def ler_riscos(caminho_kri: Path) -> dict[int, float]:
    """Mapa {ID_Risco: nota} lido das 4 abas de risco da planilha de KRI.

    A identificação das abas de risco é feita pelo CONTEÚDO (a aba tem a
    linha de cabeçalho "Fator de Risco" e uma coluna "Nota Final"/"EPS
    (KRI)"), independente do nome exato da aba. O ID do risco vem do número
    no início do nome da aba ("1...", "2.", "3 -", ...); quando não há
    número, os riscos são numerados na ordem em que aparecem na planilha.
    """
    wb = openpyxl.load_workbook(caminho_kri, data_only=True)

    abas_risco = [ws for ws in wb.worksheets if _localizar_coluna(ws) is not None]
    if not abas_risco:
        raise ValueError(
            "Nenhuma aba de risco encontrada na planilha de KRI.\n"
            f"  Abas disponíveis: {wb.sheetnames}\n"
            "  (cada aba de risco deve ter a linha 'Fator de Risco' e a "
            "coluna 'Nota Final' ou 'EPS (KRI)')."
        )

    riscos: dict[int, float] = {}
    usados: set[int] = set()
    sem_id = []
    for ws in abas_risco:
        m = re.match(r"\s*([1-4])\b", ws.title)  # número no início do nome
        if m and int(m.group(1)) not in usados:
            id_risco = int(m.group(1))
            riscos[id_risco] = extrair_nota(ws)
            usados.add(id_risco)
        else:
            sem_id.append(ws)

    # Abas sem número no nome recebem os IDs restantes, na ordem de aparição.
    livres = [i for i in (1, 2, 3, 4) if i not in usados]
    for ws, id_risco in zip(sem_id, livres):
        riscos[id_risco] = extrair_nota(ws)
        usados.add(id_risco)

    return riscos


# --------------------------------------------------------------------------- #
# Helpers da planilha destino
# --------------------------------------------------------------------------- #
def _cabecalhos(ws) -> dict[str, int]:
    """Mapa {nome_coluna_normalizado: indice_coluna} a partir da 1ª linha."""
    return {
        _normalizar(ws.cell(row=1, column=c).value): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value is not None
    }


def _como_data(valor):
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    return None


def ler_codigos_risco(wb) -> dict[int, str]:
    """Mapa {ID_Risco: Codigo} a partir da aba Dim_Risco."""
    ws = wb[ABA_RISCO]
    cols = _cabecalhos(ws)
    c_id = cols["id_risco"]
    c_cod = cols["codigo"]
    mapa = {}
    for r in range(2, ws.max_row + 1):
        idv = ws.cell(row=r, column=c_id).value
        if _eh_numero(idv):
            mapa[int(idv)] = ws.cell(row=r, column=c_cod).value
    return mapa


# --------------------------------------------------------------------------- #
# 3/4/5. Atualizar Fato_Indicadores
# --------------------------------------------------------------------------- #
def atualizar_fato(wb, data_mes: dt.datetime, riscos: dict[int, float],
                   codigos: dict[int, str]) -> None:
    ws = wb[ABA_FATO]
    cols = _cabecalhos(ws)
    c_data = cols["data"]
    c_id = cols["id_risco"]
    c_cod = cols["codigo_risco"]
    c_real = cols["real_atingimento"]
    c_meta = cols["meta"]
    c_var = cols["variacao_pp_vs_anterior"]
    c_status = cols["status_farol"]

    alvo = data_mes.date()

    for id_risco in sorted(riscos):
        nota = riscos[id_risco]

        # Mês anterior do MESMO risco (maior data anterior ao mês alvo).
        prev_valor, prev_data = None, None
        linha_existente = None
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(row=r, column=c_id).value
            if not _eh_numero(rid) or int(rid) != id_risco:
                continue
            rdata = _como_data(ws.cell(row=r, column=c_data).value)
            if rdata is None:
                continue
            if rdata == alvo:
                linha_existente = r  # já existe -> atualiza (re-execução)
            elif rdata < alvo and (prev_data is None or rdata > prev_data):
                rv = ws.cell(row=r, column=c_real).value
                if _eh_numero(rv):
                    prev_data, prev_valor = rdata, float(rv)

        variacao = (nota - prev_valor) if prev_valor is not None else 0
        farol = status_farol(nota)

        linha = linha_existente if linha_existente else ws.max_row + 1
        ws.cell(row=linha, column=c_data, value=data_mes).number_format = "mm-dd-yy"
        ws.cell(row=linha, column=c_id, value=id_risco)
        ws.cell(row=linha, column=c_cod, value=codigos.get(id_risco))
        ws.cell(row=linha, column=c_real, value=nota).number_format = "0%"
        ws.cell(row=linha, column=c_meta, value=META_PADRAO).number_format = "0%"
        ws.cell(row=linha, column=c_var, value=variacao).number_format = "0.0"
        ws.cell(row=linha, column=c_status, value=farol)

        acao = "atualizado" if linha_existente else "inserido"
        print(f"  Risco {codigos.get(id_risco, id_risco)}: "
              f"{nota:.1%} | variação {variacao:+.1%} | {farol} ({acao})")


# --------------------------------------------------------------------------- #
# 6. Atualizar Dim_Tempo
# --------------------------------------------------------------------------- #
def atualizar_tempo(wb, data_mes: dt.datetime) -> None:
    ws = wb[ABA_TEMPO]
    cols = _cabecalhos(ws)
    c_data = cols["data"]
    c_ano = cols["ano"]
    c_mes = cols["mes"]
    c_nome = cols["mes_nome"]
    c_mesano = cols["mes_ano"]
    c_tri = cols["trimestre"]
    c_atual = cols["eh_mes_atual"]

    alvo = data_mes.date()

    # Marca todos os meses como "Não"; o mês processado vira o atual.
    linha_alvo = None
    for r in range(2, ws.max_row + 1):
        rdata = _como_data(ws.cell(row=r, column=c_data).value)
        if rdata is None:
            continue
        if rdata == alvo:
            linha_alvo = r
        ws.cell(row=r, column=c_atual, value="Não")

    if linha_alvo is None:
        linha_alvo = ws.max_row + 1

    ws.cell(row=linha_alvo, column=c_data, value=data_mes).number_format = "mm-dd-yy"
    ws.cell(row=linha_alvo, column=c_ano, value=data_mes.year)
    ws.cell(row=linha_alvo, column=c_mes, value=data_mes.month)
    ws.cell(row=linha_alvo, column=c_nome, value=MES_NOME[data_mes.month])
    ws.cell(row=linha_alvo, column=c_mesano,
            value=f"{MES_NOME[data_mes.month]}/{data_mes.year % 100:02d}")
    ws.cell(row=linha_alvo, column=c_tri, value=trimestre(alvo))
    ws.cell(row=linha_alvo, column=c_atual, value="Sim")


# --------------------------------------------------------------------------- #
# Orquestração
# --------------------------------------------------------------------------- #
def processar(pasta: Path, destino: Path, ano: int, backup: bool = True) -> None:
    print(f"Pasta de KRIs : {pasta}")
    caminho_kri, mes = localizar_planilha_recente(pasta)
    data_mes = fim_do_mes(ano, mes)
    print(f"Planilha mais recente: {caminho_kri.name}  ->  "
          f"{MES_NOME[mes]}/{ano} ({data_mes.date()})")

    riscos = ler_riscos(caminho_kri)

    if not destino.exists():
        raise FileNotFoundError(f"Planilha destino não encontrada: {destino}")

    if backup:
        bkp = destino.with_name(
            f"{destino.stem}_backup_"
            f"{dt.datetime.now():%Y%m%d_%H%M%S}{destino.suffix}"
        )
        shutil.copy2(destino, bkp)
        print(f"Backup criado : {bkp.name}")

    wb = openpyxl.load_workbook(destino)
    codigos = ler_codigos_risco(wb)

    print(f"Atualizando '{ABA_FATO}':")
    atualizar_fato(wb, data_mes, riscos, codigos)

    print(f"Atualizando '{ABA_TEMPO}'... (mês atual = {MES_NOME[mes]}/{ano})")
    atualizar_tempo(wb, data_mes)

    wb.save(destino)
    print(f"Base atualizada com sucesso: {destino}")


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Atualiza a base do dashboard a partir da planilha de KRI mais recente."
    )
    parser.add_argument("--pasta", default=PASTA_KRI_PADRAO,
                        help=f"Pasta com as planilhas de KRI (padrão: {PASTA_KRI_PADRAO})")
    parser.add_argument("--destino", default=DESTINO_PADRAO,
                        help=f"Planilha destino do dashboard (padrão: {DESTINO_PADRAO})")
    parser.add_argument("--ano", type=int, default=dt.date.today().year,
                        help="Ano de referência (padrão: ano atual)")
    parser.add_argument("--sem-backup", action="store_true",
                        help="Não criar cópia de backup da planilha destino")
    args = parser.parse_args(argv)

    try:
        processar(Path(args.pasta), Path(args.destino), args.ano,
                  backup=not args.sem_backup)
    except Exception as exc:  # noqa: BLE001
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
